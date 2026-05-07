import openpyxl

wb2 = openpyxl.load_workbook(r'C:\Users\rjsx1\Downloads\lps_system\Part_Machine_Mapping_Final.xlsx', data_only=True)
print("Sheets:", wb2.sheetnames)
ws2 = wb2.active
print(f"Dimensions: {ws2.dimensions}")
print("First 15 rows:")
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i >= 15:
        break
    print(f"  Row {i+1}: {row}")
