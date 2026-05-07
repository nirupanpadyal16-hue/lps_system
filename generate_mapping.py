"""
Generate Part_Machine_Mapping_Final.xlsx from LPS G CHART APRIL'2026.xlsx Capacity sheet.

Rules:
- A part is mapped to a machine if the stroke-count column for that machine has a positive number.
- Multi-machine parts: all machines with positive stroke counts are included.
- Also include machine from M/C column (col 23) if it's a known machine string.
- Use SAP Part Number (col 4) as the identifier, matching existing file format.
- Output: ID, SAP Part Number, Assigned Machines (comma-separated), sequence_order per row.
"""

import openpyxl
import glob
from datetime import datetime
from collections import defaultdict

files = glob.glob(r'C:\Users\rjsx1\Downloads\lps_system\LPS G CHART*')
filepath = [f for f in files if not f.startswith('~')][0]
print(f"Reading: {filepath}")

wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb['Capacity']

# 0-based index -> machine name (these are the VALUE columns for each machine)
# Header row2: Col9='Blanking'(idx8), Col10=value(idx9), Col11='320T'(idx10), Col12=value(idx11),
#              Col13='200T'(idx12), Col14=value(idx13), Col15='160T'(idx14), Col16=value(idx15),
#              Col17='110T'(idx16), Col18=value(idx17), Col19='80T'(idx18), Col20=value(idx19),
#              Col21='63T'(idx20)=value, Col22='Total Strokes Req'
# Wait: Col21 header is '63T', and Col21 data = the value for 63T directly?
# From Row3: Col21=0 (idx20) - so idx20 IS the 63T value column
# And '80T' value = Col20 (idx19), '110T' value = Col18 (idx17)... etc.

# CONFIRMED mapping (0-based index of VALUE column):
STROKE_VALUE_COLS = {
    9:  'Blanking',
    11: '320T',
    13: '200T',
    15: '160T',
    17: '110T',
    19: '80T',
    20: '63T',    # Col21 is the 63T value (the header '63T' IS the label)
}

VALID_MC_STRINGS = {'Blanking', '320T', '200T', '160T', '110T', '80T', '63T', '300T'}

# Group mappings by SAP part number
part_mappings = defaultdict(lambda: {'part_description': '', 'machines': []})

row_count = 0
skipped = 0

for row in ws.iter_rows(min_row=3, values_only=True):
    sr_no = row[0]
    if sr_no is None or not isinstance(sr_no, (int, float)):
        skipped += 1
        continue

    sap_part = str(row[3]).strip() if row[3] else None   # Col4 = SAP Part Number
    part_desc = str(row[4]).strip() if row[4] else ''     # Col5 = Part Description
    mc_col = row[22]                                       # Col23 = M/C

    if not sap_part or sap_part == 'None':
        skipped += 1
        continue

    machines_for_part = []

    # Collect machines from stroke value columns
    for col_idx, machine_name in STROKE_VALUE_COLS.items():
        val = row[col_idx]
        if val is not None and isinstance(val, (int, float)) and val > 0:
            if machine_name not in machines_for_part:
                machines_for_part.append(machine_name)

    # Include M/C column if it's a recognized machine name not already captured
    if mc_col and isinstance(mc_col, str):
        mc_clean = mc_col.strip()
        if mc_clean in VALID_MC_STRINGS and mc_clean not in machines_for_part:
            machines_for_part.append(mc_clean)

    if machines_for_part:
        if sap_part not in part_mappings or not part_mappings[sap_part]['machines']:
            part_mappings[sap_part]['part_description'] = part_desc
            part_mappings[sap_part]['machines'] = machines_for_part
        else:
            # Merge if part appears multiple times (shouldn't normally happen)
            for m in machines_for_part:
                if m not in part_mappings[sap_part]['machines']:
                    part_mappings[sap_part]['machines'].append(m)

    row_count += 1

print(f"Processed {row_count} data rows, skipped {skipped}")
print(f"Unique SAP parts with machine mappings: {len(part_mappings)}")

# Count multi-machine parts
multi = {k: v for k, v in part_mappings.items() if len(v['machines']) > 1}
print(f"Parts on multiple machines: {len(multi)}")
print("\nSample multi-machine parts:")
for part, data in list(multi.items())[:15]:
    print(f"  {part}: {data['machines']}")

# ---- Write new Excel ----
out_path = r'C:\Users\rjsx1\Downloads\lps_system\Part_Machine_Mapping_Corrected.xlsx'
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Sheet1'

# Header
ws_out.append(['ID', 'SAP Part Number', 'Assigned Machines', 'Created At'])

# Style header
from openpyxl.styles import Font, PatternFill, Alignment
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
for cell in ws_out[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
row_id = 1

for sap_part, data in part_mappings.items():
    machines_str = ', '.join(data['machines'])
    ws_out.append([row_id, sap_part, machines_str, now_str])
    row_id += 1

# Auto-fit columns
ws_out.column_dimensions['A'].width = 8
ws_out.column_dimensions['B'].width = 30
ws_out.column_dimensions['C'].width = 35
ws_out.column_dimensions['D'].width = 22

wb_out.save(out_path)
print(f"\n✅ Saved {row_id - 1} rows to {out_path}")
print(f"   Multi-machine parts: {len(multi)}")
print(f"   Total machine-mapping entries (if exploded): {sum(len(v['machines']) for v in part_mappings.values())}")
