"""
VALIDATION SCRIPT — Part_Machine_Mapping_Final.xlsx
Checks all parts before importing into the database.
"""
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\rjsx1\Downloads\lps_system\Part_Machine_Mapping_Final.xlsx')
ws = wb.active

VALID_MACHINES = {'Blanking', '320T', '200T', '160T', '110T', '80T', '63T', '300T'}

total = 0
blank_part = []
blank_machine = []
invalid_machines = []
duplicate_parts = []
seen_parts = {}

all_machines_used = set()

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    rid, sap_part, machines_str, created_at = row
    total += 1

    # Check blank SAP part number
    if not sap_part or str(sap_part).strip() == '':
        blank_part.append(f"Row {i}: ID={rid}")
        continue

    sap_part = str(sap_part).strip()

    # Check duplicate SAP part numbers
    if sap_part in seen_parts:
        duplicate_parts.append(f"Row {i}: '{sap_part}' (also at row {seen_parts[sap_part]})")
    else:
        seen_parts[sap_part] = i

    # Check blank machine
    if not machines_str or str(machines_str).strip() == '':
        blank_machine.append(f"Row {i}: SAP={sap_part}")
        continue

    # Validate each machine name
    machines = [m.strip() for m in str(machines_str).split(',')]
    for m in machines:
        all_machines_used.add(m)
        if m not in VALID_MACHINES:
            invalid_machines.append(f"Row {i}: SAP={sap_part}, unknown machine='{m}'")

print("=" * 60)
print("VALIDATION REPORT — Part_Machine_Mapping_Final.xlsx")
print("=" * 60)
print(f"Total data rows: {total}")
print(f"Unique SAP parts: {len(seen_parts)}")
print()
print(f"Blank SAP part numbers: {len(blank_part)}")
for x in blank_part: print(f"  {x}")

print(f"\nBlank machine values: {len(blank_machine)}")
for x in blank_machine: print(f"  {x}")

print(f"\nDuplicate SAP parts: {len(duplicate_parts)}")
for x in duplicate_parts: print(f"  {x}")

print(f"\nInvalid machine names: {len(invalid_machines)}")
for x in invalid_machines: print(f"  {x}")

print(f"\nAll unique machine values used: {sorted(all_machines_used)}")

# Machine distribution
from collections import Counter
machine_counter = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[2]:
        for m in str(row[2]).split(','):
            machine_counter[m.strip()] += 1

print(f"\nMachine usage distribution:")
for machine, count in sorted(machine_counter.items()):
    print(f"  {machine:12s}: {count} parts")

# Multi-machine distribution
multi_counts = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[2]:
        count = len(str(row[2]).split(','))
        multi_counts[count] += 1

print(f"\nParts by number of machines:")
for num, count in sorted(multi_counts.items()):
    label = "machine" if num == 1 else "machines"
    print(f"  {num} {label}: {count} parts")

print()
if not blank_part and not blank_machine and not invalid_machines and not duplicate_parts:
    print("ALL CHECKS PASSED — Ready to import!")
else:
    print("ISSUES FOUND — Fix before importing.")
print("=" * 60)
