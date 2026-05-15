from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime
import io

from app.extensions import db
from app.models import (
    User, Demand, CarModel, InventoryItem, MasterData, MaterialData,
    RMCheckRequest, Notification, PartMachineMapping, Status
)

ppc_bp = Blueprint('ppc', __name__)

# ─────────────────────────────────────────────────────────
# Helper: get current PPC Planner user
# ─────────────────────────────────────────────────────────
def get_ppc_user():
    username = get_jwt_identity()
    return User.query.filter_by(username=username).first()


def require_ppc(fn):
    """Decorator: ensure caller is PPC_Planner or Admin."""
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = get_ppc_user()
        if not user or user.role not in ('PPC_Planner', 'Admin'):
            return jsonify({"success": False, "message": "Access denied"}), 403
        return fn(*args, **kwargs)
    return wrapper


# ─────────────────────────────────────────────────────────
# DEMANDS  (same as admin — full CRUD)
# ─────────────────────────────────────────────────────────

@ppc_bp.route('/demands', methods=['GET'])
@jwt_required()
@require_ppc
def get_demands():
    demands = Demand.query.filter_by(is_deleted=False).order_by(Demand.created_at.desc()).all()
    return jsonify({"success": True, "data": [d.to_dict() for d in demands]})


@ppc_bp.route('/demands', methods=['POST'])
@jwt_required()
@require_ppc
def create_demand():
    data = request.json or {}
    user = get_ppc_user()

    # Generate formatted_id
    count = Demand.query.count() + 1
    formatted_id = f"DEM-{count:03d}"
    while Demand.query.filter_by(formatted_id=formatted_id).first():
        count += 1
        formatted_id = f"DEM-{count:03d}"

    demand = Demand(
        formatted_id=formatted_id,
        model_id=data.get('model_id'),
        model_name=data.get('model_name', ''),
        quantity=data.get('quantity', 0),
        start_date=data.get('start_date'),
        end_date=data.get('end_date'),
        status=Status.PENDING,
        line=data.get('line'),
        manager=user.name,
        customer=data.get('customer', ''),
        company=data.get('company', ''),
    )
    db.session.add(demand)
    db.session.commit()
    return jsonify({"success": True, "data": demand.to_dict()}), 201


@ppc_bp.route('/demands/<int:demand_id>', methods=['GET'])
@jwt_required()
@require_ppc
def get_demand(demand_id):
    demand = Demand.query.get_or_404(demand_id)
    return jsonify({"success": True, "data": demand.to_dict()})


@ppc_bp.route('/demands/<int:demand_id>', methods=['PUT'])
@jwt_required()
@require_ppc
def update_demand(demand_id):
    demand = Demand.query.get_or_404(demand_id)
    data = request.json or {}
    for field in ('model_name', 'quantity', 'start_date', 'end_date', 'customer', 'company', 'line', 'status'):
        if field in data:
            setattr(demand, field, data[field])
    db.session.commit()
    return jsonify({"success": True, "data": demand.to_dict()})


@ppc_bp.route('/demands/<int:demand_id>', methods=['DELETE'])
@jwt_required()
@require_ppc
def delete_demand(demand_id):
    """
    PPC Planner soft-deletes a demand. Data is preserved in DB as backup.
    Only the is_deleted flag is set — no data is permanently removed.
    """
    demand = Demand.query.get(demand_id)
    if not demand:
        return jsonify({"success": False, "message": "Demand not found"}), 404

    demand.is_deleted = True
    demand.deleted_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"success": True, "message": "Demand removed from view. Data has been backed up in the database."})


@ppc_bp.route('/demands/<int:demand_id>/send-to-dispatch', methods=['POST'])
@jwt_required()
@require_ppc
def send_demand_to_dispatch(demand_id):
    """
    PPC Planner / Admin authorizes a completed demand for Store Keeper dispatch.
    Demand must be in PRODUCTION_DONE status.
    Sets demand status → AWAITING_STORE_DISPATCH and notifies all Store Keepers.
    """
    user = get_ppc_user()
    demand = Demand.query.get(demand_id)
    if not demand:
        return jsonify({"success": False, "message": "Demand not found"}), 404

    if demand.status not in ('PRODUCTION_DONE',):
        return jsonify({
            "success": False,
            "message": f"Demand must be in PRODUCTION_DONE status to send to dispatch. Current status: {demand.status}"
        }), 400

    demand.status = 'AWAITING_STORE_DISPATCH'
    db.session.flush()

    # Notify all Store Keepers
    store_keepers = User.query.filter_by(role='Store_Keeper', is_active=True).all()
    for sk in store_keepers:
        Notification.send(
            user_id=sk.id,
            title="Model Ready for Dispatch",
            message=f"{user.name} authorized dispatch for {demand.model_name} ({demand.formatted_id}). Please process delivery.",
            notification_type='DISPATCH_AUTHORIZED',
            demand_id=demand.id
        )

    db.session.commit()
    return jsonify({"success": True, "data": demand.to_dict()})



# ─────────────────────────────────────────────────────────
# INVENTORY  (view + Excel stock upload)
# ─────────────────────────────────────────────────────────

@ppc_bp.route('/inventory', methods=['GET'])
@jwt_required()
@require_ppc
def get_inventory():
    demand_id = request.args.get('demand_id', type=int)
    query = InventoryItem.query
    if demand_id:
        query = query.filter_by(demand_id=demand_id)
    items = query.order_by(InventoryItem.serial_number).all()
    return jsonify({"success": True, "data": [i.to_dict() for i in items]})


@ppc_bp.route('/inventory/<int:item_id>', methods=['DELETE'])
@jwt_required()
@require_ppc
def delete_inventory_item(item_id):
    """
    PPC Planner deletes a single inventory item from a demand.
    """
    from app.models import (
        InventoryItem, MachineProductionEntry, DispatchRecord, 
        RMCheckRequest, PartShortageRequest, ShortageDailyEntry
    )
    item = InventoryItem.query.get(item_id)
    if not item:
        return jsonify({"success": False, "message": "Inventory item not found"}), 404
    
    # Cleanup associated records
    MachineProductionEntry.query.filter_by(inventory_item_id=item_id).delete()
    DispatchRecord.query.filter_by(inventory_item_id=item_id).delete()
    RMCheckRequest.query.filter_by(inventory_item_id=item_id).delete()
    
    psrs = PartShortageRequest.query.filter_by(inventory_item_id=item_id).all()
    psr_ids = [p.id for p in psrs]
    if psr_ids:
        ShortageDailyEntry.query.filter(ShortageDailyEntry.shortage_request_id.in_(psr_ids)).delete(synchronize_session=False)
        PartShortageRequest.query.filter(PartShortageRequest.id.in_(psr_ids)).delete(synchronize_session=False)

    db.session.delete(item)
    db.session.commit()
    return jsonify({"success": True, "message": "Inventory item and associated data deleted"})


@ppc_bp.route('/inventory/upload-stock', methods=['POST'])
@jwt_required()
@require_ppc
def upload_stock_excel():
    """
    Excel upload to set initial_stock for InventoryItems.
    Expected columns: 'SAP PART NUMBER' (or 'sap_part_number'), 'STOCK' (or 'stock' / 'current_stock')
    Matches by sap_part_number. Updates initial_stock only (no overwrite of produced_qty).
    """
    try:
        import openpyxl
    except ImportError:
        return jsonify({"success": False, "message": "openpyxl not installed. Run: pip install openpyxl"}), 500

    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file uploaded"}), 400

    demand_id = request.form.get('demand_id', type=int)
    file = request.files['file']

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().upper() if c.value else '' for c in ws[1]]

        # Flexible header detection
        part_col = next((i for i, h in enumerate(headers) if 'SAP' in h and 'PART' in h), None)
        if part_col is None:
            part_col = next((i for i, h in enumerate(headers) if 'PART' in h and 'NUMBER' in h), None)
        stock_col = next((i for i, h in enumerate(headers) if 'STOCK' in h), None)

        if part_col is None or stock_col is None:
            return jsonify({
                "success": False,
                "message": f"Could not find required columns. Found: {headers}. Need: SAP PART NUMBER, STOCK"
            }), 400

        updated = 0
        not_found = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            sap_part = str(row[part_col]).strip() if row[part_col] else None
            stock_val = row[stock_col]
            if not sap_part or sap_part == 'None':
                continue
            try:
                stock_val = float(stock_val) if stock_val is not None else 0.0
            except (ValueError, TypeError):
                stock_val = 0.0

            # Query: match by sap_part_number (and demand_id if provided)
            q = InventoryItem.query.filter_by(sap_part_number=sap_part)
            if demand_id:
                q = q.filter_by(demand_id=demand_id)
            items = q.all()

            if items:
                for item in items:
                    item.initial_stock = stock_val
                    item.sync_current_stock()
                    # Re-evaluate status
                    if item.effective_stock >= item.demand_quantity:
                        item.status = 'SUFFICIENT'
                    else:
                        item.status = 'SHORTAGE'
                updated += len(items)
            else:
                not_found.append(sap_part)

        db.session.commit()
        return jsonify({
            "success": True,
            "message": f"Updated {updated} inventory items",
            "not_found": not_found[:20]  # Return first 20 not found for debugging
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────
# MACHINE REGISTRY — RM Check (PPC Planner view)
# Shows only parts linked to active demands
# ─────────────────────────────────────────────────────────

@ppc_bp.route('/machine-registry', methods=['GET'])
@jwt_required()
@require_ppc
def get_machine_registry():


    """
    Returns InventoryItem rows for active demands, enriched with:
    - Machine info from PartMachineMapping
    - Material data from MaterialData / MasterData
    - Demand formatted_id for identification
    - RM check request status
    """
    demand_param = request.args.get('demand_id')
    demand_id = None
    if demand_param:
        if str(demand_param).isdigit():
            demand_id = int(demand_param)
        else:
            d_obj = Demand.query.filter_by(formatted_id=demand_param).first()
            if d_obj: demand_id = d_obj.id


    # Primary filter: Only show items that have an active shortage request workflow
    query = InventoryItem.query.filter(
        InventoryItem.status.in_(['WAITING_RM_APPROVAL', 'PENDING_DEO', 'IN_PRODUCTION'])
    )

    if demand_id:
        # If user explicitly selected a demand, show its shortages regardless of general active status
        query = query.filter_by(demand_id=demand_id)
    else:
        # Active demand IDs
        active_statuses = ['PENDING', 'RM_CHECK', 'RM_SENT', 'RM_ACCEPTED', 'IN_PRODUCTION', 'IN_PROGRESS', 'IN PROGRESS', 'ACTIVE']
        active_demand_ids = db.session.query(Demand.id).filter(
            Demand.status.in_(active_statuses)
        ).all()
        active_ids = [d[0] for d in active_demand_ids]
        
        if not active_ids:
            return jsonify({"success": True, "data": []})
            
        # Otherwise, show shortages for all active demands
        query = query.filter(InventoryItem.demand_id.in_(active_ids))
    
    items = query.order_by(InventoryItem.demand_id, InventoryItem.serial_number).all()

    # Check if ANY inventory items exist for this demand (regardless of stock)
    any_exists = InventoryItem.query.filter_by(demand_id=demand_id).first() if demand_id else None

    # Fallback: If demand_id provided but NO inventory records exist at all, show MasterData parts
    if demand_id and not any_exists:
        demand = Demand.query.get(demand_id)
        if demand:
            # Use model_name from demand, or fallback to the name of the associated CarModel
            search_name = demand.model_name
            if not search_name and demand.model:
                search_name = demand.model.name
            
            if search_name:
                # Robust matching: Try exact match first, then ilike
                master_parts = MasterData.query.filter(
                    (db.func.lower(MasterData.model) == search_name.lower()) |
                    (MasterData.model.ilike(f"%{search_name}%"))
                ).all()
                
                result = []
                for idx, part in enumerate(master_parts, start=1):
                    mapping = PartMachineMapping.query.filter_by(sap_part_number=part.sap_part_number).first()
                    result.append({
                        "id": 0, # Virtual ID for frontend to recognize as new/unseeded
                        "demand_id": demand.id,
                        "demand_formatted_id": demand.formatted_id,
                        "serial_number": idx,
                        "vehicle_name": search_name,
                        "sap_part_number": part.sap_part_number,
                        "part_number": part.part_number,
                        "assembly_number": part.assembly_number,
                        "part_description": part.description,
                        "current_stock": 0,
                        "demand_quantity": demand.quantity,
                        "shortage_quantity": demand.quantity,
                        "machine_group": mapping.machine.replace(', ', ' → ') if mapping and mapping.machine else "UNASSIGNED",
                        "rm_status": "NOT_SEEDED",
                        "master_material_data": part.material_rel.to_dict() if part.material_rel else {},
                        "rm_request": None
                    })
                return jsonify({"success": True, "data": result})


    result = []
    for item in items:
        d = item.to_dict()
        # Ensure demand_formatted_id is present
        if not d.get('demand_formatted_id'):
            demand = Demand.query.get(item.demand_id)
            d['demand_formatted_id'] = demand.formatted_id if demand else None

        # Machine info
        mapping = PartMachineMapping.query.filter_by(sap_part_number=item.sap_part_number).first()
        d['machine_group'] = mapping.machine.replace(', ', ' → ') if mapping and mapping.machine else None

        # Material data from MasterData
        master = MasterData.query.filter_by(sap_part_number=item.sap_part_number).first()
        if master:
            d['part_number'] = master.part_number
            d['assembly_number'] = master.assembly_number
            if master.material_rel:
                d['master_material_data'] = master.material_rel.to_dict()
            else:
                d['master_material_data'] = {}
        else:
            d['master_material_data'] = {}

        # Latest RMCheckRequest for this item
        latest_rm = RMCheckRequest.query.filter_by(
            inventory_item_id=item.id
        ).order_by(RMCheckRequest.created_at.desc()).first()
        d['rm_request'] = latest_rm.to_dict() if latest_rm else None

        result.append(d)

    return jsonify({"success": True, "data": result})


@ppc_bp.route('/machine-registry/<int:item_id>/rm-data', methods=['GET'])
@jwt_required()
@require_ppc
def get_rm_data(item_id):
    """Get RM data for a single inventory item (master data + any existing RMCheckRequest)."""
    item = InventoryItem.query.get_or_404(item_id)
    master = MasterData.query.filter_by(sap_part_number=item.sap_part_number).first()
    master_mat = master.material_rel.to_dict() if master and master.material_rel else {}

    latest_rm = RMCheckRequest.query.filter_by(
        inventory_item_id=item_id
    ).order_by(RMCheckRequest.created_at.desc()).first()

    return jsonify({
        "success": True,
        "inventory_item": item.to_dict(),
        "master_material_data": master_mat,
        "rm_request": latest_rm.to_dict() if latest_rm else None
    })


@ppc_bp.route('/machine-registry/<int:item_id>/submit-rm', methods=['POST'])
@jwt_required()
@require_ppc
def submit_rm_to_store_keeper(item_id):
    """
    PPC Planner edits RM data for a part and submits to Store Keeper.
    Creates a new RMCheckRequest — does NOT modify MasterData.
    """
    data = request.json or {}
    user = get_ppc_user()
    
    # Handle virtual items (item_id=0) or items not yet in InventoryItem table
    if item_id == 0:
        demand_id = data.get('demand_id')
        sap_part_number = data.get('sap_part_number')
        if not demand_id or not sap_part_number:
            return jsonify({"success": False, "message": "Missing demand_id or sap_part_number for virtual item"}), 400
        
        # Check if it was seeded while we were editing
        item = InventoryItem.query.filter_by(demand_id=demand_id, sap_part_number=sap_part_number).first()
        if not item:
            # Create the InventoryItem on the fly
            demand = Demand.query.get_or_404(demand_id)
            master = MasterData.query.filter_by(sap_part_number=sap_part_number).first_or_404()
            item = InventoryItem(
                demand_id=demand_id,
                sap_part_number=sap_part_number,
                part_description=master.description,
                demand_quantity=demand.quantity,
                shortage_quantity=demand.quantity,
                status='PENDING',
                rm_status='RM_SUBMITTED'
            )
            db.session.add(item)
            db.session.flush()
    else:
        item = InventoryItem.query.get_or_404(item_id)

    # Generate RMR formatted_id
    count = RMCheckRequest.query.count() + 1
    formatted_id = f"RMR-{count:04d}"
    while RMCheckRequest.query.filter_by(formatted_id=formatted_id).first():
        count += 1
        formatted_id = f"RMR-{count:04d}"

    rm_req = RMCheckRequest(
        formatted_id=formatted_id,
        inventory_item_id=item.id,
        demand_id=item.demand_id,
        ppc_planner_id=user.id,
        rm_thk_mm=data.get('rm_thk_mm'),
        sheet_width=data.get('sheet_width'),
        sheet_length=data.get('sheet_length'),
        no_of_comp_per_sheet=data.get('no_of_comp_per_sheet'),
        rm_size=data.get('rm_size'),
        rm_grade=data.get('rm_grade'),
        act_rm_sizes=data.get('act_rm_sizes'),
        ppc_notes=data.get('ppc_notes'),
        status='SUBMITTED',
        submitted_at=datetime.utcnow(),
    )
    db.session.add(rm_req)

    # Update inventory item rm_status
    item.rm_status = 'RM_SUBMITTED'

    db.session.commit()

    # Send notification to all Store Keepers
    store_keepers = User.query.filter_by(role='Store_Keeper', is_active=True).all()
    for sk in store_keepers:
        Notification.send(
            user_id=sk.id,
            title="New RM Sheet Submitted",
            message=f"PPC Planner {user.name} submitted RM data for part {item.sap_part_number} (Demand: {rm_req.formatted_id})",
            notification_type='RM_SUBMITTED',
            demand_id=item.demand_id,
            rm_request_id=rm_req.id
        )
    db.session.commit()

    return jsonify({"success": True, "data": rm_req.to_dict()}), 201


@ppc_bp.route('/rm-requests', methods=['GET'])
@jwt_required()
@require_ppc
def get_my_rm_requests():
    """List all RMCheckRequests submitted by this PPC Planner."""
    user = get_ppc_user()
    reqs = RMCheckRequest.query.filter_by(ppc_planner_id=user.id).order_by(
        RMCheckRequest.created_at.desc()
    ).all()
    return jsonify({"success": True, "data": [r.to_dict() for r in reqs]})


# ─────────────────────────────────────────────────────────
# NOTIFICATIONS (read own)
# ─────────────────────────────────────────────────────────

@ppc_bp.route('/notifications', methods=['GET'])
@jwt_required()
@require_ppc
def get_notifications():
    user = get_ppc_user()
    notifs = Notification.query.filter_by(user_id=user.id).order_by(
        Notification.created_at.desc()
    ).limit(50).all()
    unread = Notification.query.filter_by(user_id=user.id, is_read=False).count()
    return jsonify({"success": True, "data": [n.to_dict() for n in notifs], "unread_count": unread})


@ppc_bp.route('/notifications/<int:notif_id>/read', methods=['POST'])
@jwt_required()
@require_ppc
def mark_notification_read(notif_id):
    user = get_ppc_user()
    notif = Notification.query.filter_by(id=notif_id, user_id=user.id).first_or_404()
    notif.is_read = True
    notif.read_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"success": True})
